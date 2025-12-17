"""
Excel到IDR转换模块 - 极简版

将Excel数据转换为标准IDR格式，提取图片资产。
"""

import json
import ast
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Union, Tuple, Optional
import re
import operator

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# 简化日志
def log(msg: str) -> None:
    """简化日志输出"""
    print(f"[EXCEL2JSON] {msg}")


# 全局缓存用于工作簿实例
_workbook_cache: Dict[str, Any] = {}


class ExcelWorkbookCache:
    """Excel工作簿缓存类，避免重复加载"""

    def __init__(self, excel_path: Union[str, Path]):
        self.excel_path = str(excel_path)
        self._workbook = None
        self._data_workbook = None  # 用于获取计算值的工作簿
        self._cache = {}

    @property
    def workbook(self):
        """延迟加载公式工作簿"""
        if self._workbook is None:
            try:
                self._workbook = load_workbook(self.excel_path, data_only=False)
                log(f"  加载工作簿: {Path(self.excel_path).name}")
            except Exception as e:
                log(f"  工作簿加载失败: {e}")
                raise
        return self._workbook

    @property
    def data_workbook(self):
        """延迟加载值工作簿"""
        if self._data_workbook is None:
            try:
                self._data_workbook = load_workbook(self.excel_path, data_only=True)
            except Exception as e:
                log(f"  值工作簿加载失败: {e}")
                # 如果无法加载值工作簿，返回公式工作簿
                self._data_workbook = self.workbook
        return self._data_workbook

    def get_worksheet(self, sheet_name: str) -> Worksheet:
        """获取工作表"""
        return self.workbook[sheet_name]

    def get_data_worksheet(self, sheet_name: str) -> Worksheet:
        """获取值工作表"""
        return self.data_workbook[sheet_name]

    def get_cache_value(self, key: str) -> Optional[Union[float, int]]:
        """获取缓存值"""
        return self._cache.get(key)

    def set_cache_value(self, key: str, value: Union[float, int]) -> None:
        """设置缓存值"""
        self._cache[key] = value


def safe_eval(expression: str) -> Optional[Union[float, int]]:
    """
    安全的数学表达式计算

    Args:
        expression: 要计算的数学表达式

    Returns:
        计算结果或None（如果表达式不安全）
    """
    try:
        # 使用AST解析确保只允许数学运算
        node = ast.parse(expression, mode='eval')

        # 验证AST节点类型 - 只允许安全的数学运算
        allowed_nodes = (
            ast.Expression, ast.BinOp, ast.UnaryOp, ast.Constant,
            ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Pow, ast.Mod,
            ast.USub, ast.UAdd
        )

        # 为了兼容旧版本Python，也允许ast.Num
        try:
            ast.Num  # 检查是否存在
            allowed_nodes = allowed_nodes + (ast.Num,)
        except AttributeError:
            pass  # Python 3.8+没有ast.Num

        allowed_operators = {
            ast.Add: operator.add,
            ast.Sub: operator.sub,
            ast.Mult: operator.mul,
            ast.Div: operator.truediv,
            ast.Pow: operator.pow,
            ast.Mod: operator.mod,
        }

        def eval_node(node):
            if isinstance(node, ast.Constant):
                if isinstance(node.value, (int, float)):
                    return node.value
                return None
            elif hasattr(ast, 'Num') and isinstance(node, ast.Num):  # Python < 3.8
                return node.n
            elif isinstance(node, ast.BinOp):
                left = eval_node(node.left)
                right = eval_node(node.right)
                if left is not None and right is not None and type(node.op) in allowed_operators:
                    return allowed_operators[type(node.op)](left, right)
                return None
            elif isinstance(node, ast.UnaryOp):
                operand = eval_node(node.operand)
                if operand is not None and type(node.op) in allowed_operators:
                    return allowed_operators[type(node.op)](operand)
                return None
            else:
                return None

        result = eval_node(node.body)
        return result

    except (ValueError, SyntaxError, TypeError, ZeroDivisionError):
        return None
    except Exception:
        return None


def calculate_simple_formula(
    formula_str: str,
    worksheet,
    row: int,
    col: int,
    cache: Optional[Dict[str, Union[float, int]]] = None,
    workbook_cache: Optional[ExcelWorkbookCache] = None
) -> Optional[Union[float, int]]:
    """
    计算简单的Excel公式
    支持基本的算术运算：+、-、*、/
    支持单元格引用和递归计算
    """
    try:
        # 移除开头的等号
        if formula_str.startswith('='):
            formula_str = formula_str[1:]

        # 处理简单的单元格引用公式，如 =B2*C2
        # 使用正则表达式找到所有单元格引用
        cell_refs = re.findall(r'([A-Z]+\d+)', formula_str)

        # 替换单元格引用为实际值
        calculated_formula = formula_str
        for ref in cell_refs:
            try:
                ref_cell = worksheet[ref]
                ref_value = ref_cell.value

                # 首先检查缓存中是否有这个单元格的值
                sheet_name = worksheet.title
                ref_key = f"{sheet_name}_{ref_cell.row}_{ref_cell.column}"

                if cache and ref_key in cache:
                    ref_value = cache[ref_key]
                # 如果缓存没有，尝试获取单元格的计算值
                else:
                    # 如果是公式单元格，需要获取计算后的值
                    if ref_cell.data_type == 'f':
                        # 递归计算引用的公式
                        ref_value = calculate_simple_formula(str(ref_cell.value), worksheet,
                                                          ref_cell.row, ref_cell.column, cache, workbook_cache)
                    # 如果是普通单元格但值为空，尝试读取显示值
                    elif ref_value is None:
                        try:
                            # 从缓存的值工作簿获取计算值
                            data_sheet = workbook_cache.get_data_worksheet(sheet_name)
                            ref_value = data_sheet[ref].value
                        except Exception:
                            ref_value = None

                    if cache is not None and ref_value is not None:
                        cache[ref_key] = ref_value

                if ref_value is not None and not isinstance(ref_value, str):
                    calculated_formula = calculated_formula.replace(ref, str(ref_value))
                else:
                    # 如果引用的单元格没有有效值，用0代替
                    log(f"    引用单元格 {ref} 无有效值: {ref_value}，使用0代替")
                    calculated_formula = calculated_formula.replace(ref, '0')
            except Exception as e:
                log(f"    处理引用单元格 {ref} 失败: {e}")
                # 不要立即返回None，继续尝试其他引用单元格
                continue

        # 替换Excel函数为Python等价函数（简单示例）
        calculated_formula = calculated_formula.replace('SUM(', '(')

        # 使用安全的表达式计算
        result = safe_eval(calculated_formula)

        # 如果所有引用都已替换，但仍有常数表达式，直接计算
        if result is None:
            try:
                # 检查是否只剩下常数和运算符
                if re.match(r'^[\d+\-*/.()\s]+$', calculated_formula):
                    result = safe_eval(calculated_formula)
            except Exception:
                pass

        return result

    except Exception as e:
        log(f"    公式计算失败: {formula_str} -> {e}")
        return None


def extract_images(
    excel_path: Union[str, Path],
    assets_dir: Union[str, Path]
) -> List[Dict[str, Any]]:
    """
    提取Excel中的图片资产

    Args:
        excel_path: Excel文件路径
        assets_dir: 图片保存目录

    Returns:
        图片资产列表
    """
    assets_dir = Path(assets_dir)
    assets_dir.mkdir(parents=True, exist_ok=True)

    try:
        log(f"提取图片: {Path(excel_path).name}")
        workbook = load_workbook(excel_path)
        assets: List[Dict[str, Any]] = []
        image_id = 1

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # 尝试使用公共API，失败后使用私有API
            images = getattr(worksheet, 'images', None)
            if images is None:
                images = getattr(worksheet, '_images', [])

            for image in images:
                try:
                    # 生成资产信息
                    asset_id = f"img_{image_id:03d}"
                    image_id += 1

                    # 获取位置信息
                    try:
                        row = image.anchor._from.row + 1
                        col = image.anchor._from.col + 1
                    except (AttributeError, TypeError):
                        row = col = 1

                    # 保存图片文件
                    filename = f"{asset_id}.png"
                    image_path = assets_dir / filename

                    # 获取图片数据
                    img_data = None
                    if hasattr(image, 'bytes'):
                        img_data = image.bytes
                    elif hasattr(image, '_data'):
                        try:
                            img_data = image._data()
                        except (AttributeError, TypeError):
                            pass

                    if img_data is not None:
                        with open(image_path, 'wb') as f:
                            f.write(img_data)

                        # 获取尺寸
                        try:
                            width = float(image.width or 100)
                            height = float(image.height or 100)
                        except (AttributeError, TypeError, ValueError):
                            width = height = 100.0

                        assets.append({
                            "asset_id": asset_id,
                            "filename": filename,
                            "sheet_name": sheet_name,
                            "row": row,
                            "col": col,
                            "width": width,
                            "height": height
                        })

                except (AttributeError, TypeError, OSError, IOError) as e:
                    log(f"  跳过图片处理错误: {e}")
                    continue

        log(f"  提取图片: {len(assets)} 个")
        return assets

    except (FileNotFoundError, PermissionError) as e:
        log(f"文件访问错误: {e}")
        return []
    except Exception as e:
        log(f"图片提取失败: {e}")
        return []


def extract_data(excel_path: Union[str, Path]) -> Tuple[List[Dict[str, Any]], List[Tuple[str, int]]]:
    """提取Excel数据 - 使用缓存优化确保公式值正确读取
    返回: (数据记录, 元数据列表[(sheet_name, row_num)])
    """
    try:
        log(f"提取数据: {Path(excel_path).name}")
        log("  正在使用openpyxl读取公式值...")

        # 使用缓存的工作簿
        workbook_cache = ExcelWorkbookCache(excel_path)
        workbook = workbook_cache.workbook

        all_records = []
        all_metadata = []  # 用于图片关联的元数据

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            log(f"  处理工作表: {sheet_name}")

            # 获取数据范围
            max_row = worksheet.max_row
            max_col = worksheet.max_column

            if max_row < 2:  # 至少需要标题行和一行数据
                continue

            # 读取标题行
            headers = []
            for col in range(1, max_col + 1):
                header_value = worksheet.cell(row=1, column=col).value
                if header_value is not None:
                    headers.append(str(header_value).strip())
                else:
                    headers.append(f"Column_{col}")

            # 读取数据行
            records = []
            for row in range(2, max_row + 1):
                # 检查是否为空行
                row_has_data = False
                record = {}

                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell_value = cell.value

                    # 处理公式单元格
                    if cell.data_type == 'f' and cell_value:
                        formula_str = str(cell_value).strip()
                        cell_key = f"{sheet_name}_{row}_{col}"

                        # 如果已经计算过，使用缓存
                        cached_value = workbook_cache.get_cache_value(cell_key)
                        if cached_value is not None:
                            record[headers[col-1]] = cached_value
                        else:
                            # 尝试计算公式
                            log(f"    发现公式单元格({row},{col}): {formula_str}")
                            calculated_value = calculate_simple_formula(
                                formula_str, worksheet, row, col, workbook_cache._cache, workbook_cache
                            )

                            if calculated_value is not None:
                                record[headers[col-1]] = calculated_value
                                workbook_cache.set_cache_value(cell_key, calculated_value)
                            else:
                                record[headers[col-1]] = f"[公式: {formula_str}]"

                        row_has_data = True
                        continue

                    # 处理普通单元格
                    if cell_value is not None:
                        row_has_data = True
                        # 清理字符串值
                        if isinstance(cell_value, str):
                            record[headers[col-1]] = cell_value.strip()
                        else:
                            record[headers[col-1]] = cell_value
                    else:
                        record[headers[col-1]] = None

                if row_has_data:
                    records.append(record)
                    # 添加元数据用于图片关联
                    all_metadata.append((sheet_name, row))

            all_records.extend(records)
            log(f"  工作表 {sheet_name}: {len(records)} 条")

        log(f"  总数据: {len(all_records)} 条")
        return all_records, all_metadata

    except Exception as e:
        log(f"数据提取失败: {e}")
        return [], []




def build_idr(excel_path: str, data: List[Dict[str, Any]], metadata: List[Tuple[str, int]], assets: List[Dict[str, Any]]) -> Dict[str, Any]:
    """构建IDR结构"""
    # 关联图片到数据记录
    clean_data = []
    for i, record in enumerate(data):
        clean_record = record.copy()  # 创建干净的记录副本

        # 查找对应的图片
        images = []
        if i < len(metadata):
            sheet_name, row_num = metadata[i]
            for asset in assets:
                if (asset['sheet_name'] == sheet_name and
                    asset['row'] == row_num):
                    images.append({
                        "asset_id": asset['asset_id'],
                        "filename": asset['filename']
                    })

        # 将图片信息添加到"图片"字段
        if images:
            clean_record['图片'] = images

        clean_data.append(clean_record)

    # 构建IDR
    return {
        "metadata": {
            "source": Path(excel_path).name,
            "created": datetime.now().isoformat(),
            "records": len(clean_data),
            "assets": len(assets)
        },
        "data": clean_data,
        "assets": assets
    }


def excel_to_idr(excel_path: Union[str, Path], output_dir: Union[str, Path]) -> str:
    """主转换函数"""
    excel_path = Path(excel_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    log(f"开始转换: {excel_path.name}")

    # 创建输出目录结构
    excel_name = excel_path.stem
    excel_output_dir = output_dir / excel_name
    excel_output_dir.mkdir(exist_ok=True)

    assets_dir = excel_output_dir / "images"

    # 提取数据和图片
    data, metadata = extract_data(excel_path)
    assets = extract_images(excel_path, assets_dir)

    # 构建IDR
    idr_data = build_idr(str(excel_path), data, metadata, assets)

    # 保存文件
    output_file = excel_output_dir / f"{excel_name}.json"

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(idr_data, f, ensure_ascii=False, indent=2)

    log(f"转换完成: {output_file.name}")
    return str(output_file)


def batch_convert(input_dir: Union[str, Path] = "data/preprocess",
                 output_dir: Union[str, Path] = "data/temp") -> Dict[str, str]:
    """批量转换"""
    input_path = Path(input_dir)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    log(f"批量转换: {input_path}")

    # 查找Excel文件
    excel_files = list(input_path.glob("*.xlsx")) + list(input_path.glob("*.xls"))

    if not excel_files:
        log("未找到Excel文件")
        return {}

    results = {}
    for excel_file in excel_files:
        try:
            output_file = excel_to_idr(excel_file, output_path)
            results[excel_file.name] = output_file
        except Exception as e:
            log(f"转换失败 {excel_file.name}: {e}")

    log(f"批量完成: {len(results)} 个文件")
    return results


def main():
    """命令行入口"""
    import argparse

    parser = argparse.ArgumentParser(description="Excel到IDR转换")
    parser.add_argument("--input", "-i", default="data/preprocess", help="输入目录")
    parser.add_argument("--output", "-o", default="data/temp", help="输出目录")
    parser.add_argument("--file", "-f", help="单个文件")

    args = parser.parse_args()

    if args.file:
        # 单文件转换
        try:
            result = excel_to_idr(args.file, args.output)
            print(f"转换成功: {result}")
        except Exception as e:
            print(f"转换失败: {e}")
            exit(1)
    else:
        # 批量转换
        results = batch_convert(args.input, args.output)
        if results:
            print(f"转换完成 {len(results)} 个文件:")
            for src, dst in results.items():
                print(f"  {src} -> {dst}")
        else:
            print("未找到可转换的文件")
            exit(1)


if __name__ == "__main__":
    main()