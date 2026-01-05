"""
Excel预处理 - 极简版

保留格式，删除无用sheet和空行
"""

import sys
from pathlib import Path
import openpyxl
import glob


# 简化日志
def log(msg):
    print(f"[PREPROCESS] {msg}")


def is_useful_sheet(worksheet):
    """判断工作表是否有用"""
    # 检查第一行是否包含标准列头
    headers = []
    for col in range(1, min(15, worksheet.max_column + 1)):  # 只检查前15列
        cell_val = worksheet.cell(1, col).value
        if cell_val and str(cell_val).strip():
            headers.append(str(cell_val).strip())

    # 必须包含的关键列头
    required_headers = [
        'FBA箱号',
        '中文品名',
        '英文品名',
        'SKU码',
        '海关编码',
        '材质（中文）',
        '材质（英文）',
        '品牌',
        '品牌类型',
        '型号'
    ]

    # 检查是否包含大部分必需的列头
    header_text = ' '.join(headers)
    matched_count = sum(1 for header in required_headers if header in header_text)

    # 如果匹配了大部分（至少10个）必需列头，则认为是有用的sheet
    return matched_count >= 10


def remove_empty_rows(worksheet):
    """删除空行（保留第1行标题）"""
    deleted_count = 0

    # 从后往前删除，避免索引变化
    for row in range(worksheet.max_row, 1, -1):  # 从最后一行到第2行
        non_empty_count = 0

        # 统计该行有效非空的列数
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row, col).value
            if cell_value and str(cell_value).strip():
                non_empty_count += 1

        # 如果有效非空列数小于10个，删除该行
        if non_empty_count < 10:
            worksheet.delete_rows(row)
            deleted_count += 1

    return deleted_count


def preprocess_excel(excel_path, output_dir="data/preprocess"):
    """预处理单个Excel文件

    Args:
        excel_path: 输入的Excel文件路径
        output_dir: 输出目录（默认: data/preprocess）

    Returns:
        成功处理的文件列表
    """
    excel_path = Path(excel_path)
    output_dir = Path(output_dir)
    log(f"处理: {excel_path.name}")

    try:
        # 第一步：加载原始工作簿
        wb = openpyxl.load_workbook(excel_path)

        # 找出有用的工作表
        useful = [name for name in wb.sheetnames
                  if is_useful_sheet(wb[name])]

        if not useful:
            log("  无有用工作表")
            return []

        # 保存结果
        output_files = []
        output_dir.mkdir(parents=True, exist_ok=True)

        if len(useful) == 1:
            # 只有一个有用sheet，直接删除其他sheet后保存
            sheet_name = useful[0]

            # 删除其他sheet
            for sheet in wb.sheetnames[:]:
                if sheet != sheet_name:
                    del wb[sheet]

            # 第二步：在唯一的sheet中删除空行
            ws = wb[sheet_name]
            deleted_rows = remove_empty_rows(ws)
            log(f"  删除空行: {deleted_rows} 行")

            output_path = output_dir / excel_path.name
            wb.save(output_path)
            output_files.append(str(output_path))
            log(f"  保存: {excel_path.name}")

        else:
            # 多个有用sheet，采用复制整个文件的方式
            import shutil

            for i, sheet_name in enumerate(useful):
                # 复制整个原始Excel文件（保持所有格式、图片等）
                temp_name = f"{excel_path.stem}_temp_{i}.xlsx"
                temp_path = output_dir / temp_name
                shutil.copy2(excel_path, temp_path)

                # 加载复制的文件
                new_wb = openpyxl.load_workbook(temp_path)

                # 删除不需要的sheet，只保留当前这一个
                for sheet in new_wb.sheetnames[:]:
                    if sheet != sheet_name:
                        del new_wb[sheet]

                # 在保留的sheet中删除空行
                ws = new_wb[sheet_name]
                deleted_rows = remove_empty_rows(ws)
                log(f"  处理工作表: {sheet_name}, 删除空行: {deleted_rows} 行")

                # 重命名为最终文件名
                final_name = f"{excel_path.stem}_{sheet_name}.xlsx"
                final_path = output_dir / final_name
                new_wb.save(final_path)

                # 删除临时文件
                temp_path.unlink()

                output_files.append(str(final_path))
                log(f"  保存: {final_name}")

        log(f"完成: {len(output_files)} 个文件")
        return output_files

    except Exception as e:
        log(f"错误: {e}")
        import traceback
        traceback.print_exc()
        return []


def main():
    """命令行入口"""
    import argparse

    parser = argparse.ArgumentParser(description="Excel预处理 - 极简版")
    parser.add_argument("--file", "-f", help="处理单个文件")
    parser.add_argument("--input", "-i", default="data/raw", help="输入目录")
    parser.add_argument("--output", "-o", default="data/preprocess", help="输出目录")

    args = parser.parse_args()

    if args.file:
        # 单文件处理
        files = preprocess_excel(args.file)
        if files:
            print(f"成功: {len(files)} 个文件")
        else:
            print("失败")
            sys.exit(1)
    else:
        # 批量处理
        excel_files = glob.glob(f"{args.input}/*.xlsx") + glob.glob(f"{args.input}/*.xls")

        if not excel_files:
            print("未找到Excel文件")
            sys.exit(1)

        print(f"处理 {len(excel_files)} 个文件")
        for excel_file in excel_files:
            preprocess_excel(excel_file)


if __name__ == "__main__":
    main()
