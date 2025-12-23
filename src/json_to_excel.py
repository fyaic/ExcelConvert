"""
JSON转Excel转换模块 - 支持新数据格式和图片处理

功能：
- 从data/transformed/读取JSON文件
- 从data/temp/读取图片
- 输出到data/output/目录
- 支持新的JSON图片字段格式
- 保持智能字段匹配和图片处理功能
- 完全遵循模板格式和样式
"""

import json
import os
import sys
import shutil
import argparse
from pathlib import Path
from typing import Any, Dict, List, Union
import logging

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

# 默认配置
DEFAULT_TEMPLATE = "data/templates/客户货物运输托运书.xlsx"
IMAGE_MAX_WIDTH = 300
IMAGE_MAX_HEIGHT = 200

# 设置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def log(msg: str) -> None:
    """简化日志输出"""
    print(f"[JSON2EXCEL] {msg}")


def load_json_data(file_path: str) -> dict:
    """加载新格式的JSON数据"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)

        # 提取数据记录
        records = json_data.get('data', [])

        # 处理新的图片字段格式
        for record in records:
            if '图片' in record and isinstance(record['图片'], list):
                record['_images'] = record['图片']

        log(f"  加载数据: {len(records)} 条记录")
        return json_data

    except Exception as e:
        log(f"JSON加载失败: {e}")
        return {'data': []}


def find_template_columns(template_path: str) -> List[str]:
    """获取模板文件的列顺序"""
    try:
        df_template = pd.read_excel(template_path, nrows=0)
        columns = list(df_template.columns)
        log(f"  模板列: {len(columns)} 个")
        return columns

    except Exception as e:
        log(f"模板读取失败: {e}")
        return []


def resolve_image_directory(json_file: str) -> str:
    """解析图片目录路径 - 完全参考old版本的逻辑"""
    # 获取JSON文件名（去掉路径和扩展名）
    json_filename = os.path.splitext(os.path.basename(json_file))[0]

    # 如果是transformed目录下的文件，提取公司名
    if "transformed" in json_file:
        # 例如: data/transformed/CDKJ/CDKJ.json -> CDKJ
        company_name = os.path.basename(os.path.dirname(json_file))
    else:
        company_name = json_filename

    image_dir = None

    # 首先尝试images子目录
    candidate_dir = os.path.join("data/temp", company_name, "images")
    if os.path.exists(candidate_dir):
        image_dir = candidate_dir
    else:
        # 尝试公司名目录
        candidate_dir = os.path.join("data/temp", company_name)
        if os.path.exists(candidate_dir):
            image_dir = candidate_dir
        else:
            # 如果没有精确匹配，尝试匹配sheet目录（多sheet情况）
            temp_dir = "data/temp"
            if os.path.exists(temp_dir):
                for subdir in os.listdir(temp_dir):
                    subdir_path = os.path.join(temp_dir, subdir)
                    if os.path.isdir(subdir_path):
                        # 检查是否包含原始文件名（对于多sheet目录）
                        if company_name in subdir or subdir.startswith(company_name.split('_sheet')[0]):
                            image_dir = subdir_path
                            break

                # 如果还没找到，尝试第一个子目录的images子目录
                if image_dir is None:
                    for item in os.listdir("data/temp"):
                        item_path = os.path.join("data/temp", item)
                        images_path = os.path.join(item_path, "images")
                        if os.path.isdir(images_path):
                            image_dir = images_path
                            break

                # 最后尝试第一个有图片的目录
                if image_dir is None:
                    for subdir in os.listdir(temp_dir):
                        subdir_path = os.path.join(temp_dir, subdir)
                        images_path = os.path.join(subdir_path, "images")

                        # 优先查找images子目录
                        if os.path.isdir(images_path):
                            for file in os.listdir(images_path):
                                if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):
                                    image_dir = images_path
                                    break
                        # 如果images子目录没有图片，查找主目录
                        elif os.path.isdir(subdir_path):
                            for file in os.listdir(subdir_path):
                                if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):
                                    image_dir = subdir_path
                                    break

                        if image_dir:
                            break

    if image_dir:
        log(f"  图片目录: {image_dir}")
    else:
        log(f"  未找到图片目录: {company_name}")

    return image_dir


def normalize_field_name(field_name: str) -> str:
    """标准化字段名，用于模糊匹配 - 完全参考old版本"""
    if not field_name:
        return ""

    import re
    # 替换所有非字母数字中文字符为空
    normalized = re.sub(r'[^\w\u4e00-\u9fff]', '', str(field_name))
    return normalized.lower()


def map_fields_to_template(record: Dict[str, Any], template_columns: List[str]) -> Dict[str, Any]:
    """智能字段映射 - 修复FBA字段匹配bug的版本 + 修复尺寸字段换行符匹配"""
    mapped_record = {}

    # 先为FBA字段建立精确映射表，避免交叉匹配
    fba_field_map = {}
    for json_field in record.keys():
        if 'FBA箱号' in json_field or ('FBA' in json_field and '箱号' in json_field):
            fba_field_map['FBA箱号'] = json_field
        elif 'FBA仓库地址' in json_field:
            fba_field_map['FBA仓库地址'] = json_field
        elif '仓库代码' in json_field and 'AMAZON' in json_field:
            fba_field_map['仓库代码_AMAZON'] = json_field

    for template_col in template_columns:
        value = None

        # 1. 直接匹配（最优先）
        if template_col in record:
            value = record[template_col]
            mapped_record[template_col] = value
            continue

        # 2. 尺寸字段特殊处理 - 处理空格和换行符的差异
        # 模板中的格式: "长\ncm", "宽\ncm", "高\ncm" (带换行符)
        # 数据中的格式: "长 cm", "宽 cm", "高 cm" (带空格)
        dimension_fields = {
            '长': ['长 cm', '长\ncm'],
            '宽': ['宽 cm', '宽\ncm'],
            '高': ['高 cm', '高\ncm']
        }
        for dim_key, variants in dimension_fields.items():
            if dim_key in template_col and 'cm' in template_col.lower():
                # 找到数据中的对应字段（不管空格还是换行符）
                for variant in variants:
                    if variant in record:
                        value = record[variant]
                        mapped_record[template_col] = value
                        break
                if value is not None:
                    break
        if value is not None:
            continue

        # 3. 处理换行符差异（通用处理）
        # 将模板列中的换行符替换为空格来匹配
        template_with_space = template_col.replace('\n', ' ')
        if template_with_space in record:
            value = record[template_with_space]
            mapped_record[template_col] = value
            continue

        # 将模板列中的空格替换为换行符来匹配
        template_with_newline = template_col.replace(' ', '\n')
        if template_with_newline in record:
            value = record[template_with_newline]
            mapped_record[template_col] = value
            continue

        # 3. FBA字段精确匹配 - 最高优先级，避免交叉匹配
        if 'FBA' in template_col:
            # FBA箱号匹配 - 必须严格匹配
            if '箱号' in template_col and 'FBA箱号' in fba_field_map:
                value = record[fba_field_map['FBA箱号']]
                mapped_record[template_col] = value
                continue
            # FBA仓库地址匹配 - 必须严格匹配
            elif '仓库' in template_col and '地址' in template_col and 'FBA仓库地址' in fba_field_map:
                value = record[fba_field_map['FBA仓库地址']]
                mapped_record[template_col] = value
                continue
            # 仓库代码匹配 - 必须严格匹配
            elif '仓库代码' in template_col and 'AMAZON' in template_col and '仓库代码_AMAZON' in fba_field_map:
                value = record[fba_field_map['仓库代码_AMAZON']]
                mapped_record[template_col] = value
                continue

        # 4. 处理引号差异的精确匹配
        quotes = ['"', '"', '"', '"', '"', '"', ''', ''']
        template_clean = template_col
        for quote in quotes:
            template_clean = template_clean.replace(quote, '')

        # 对JSON字段进行同样的清理
        for json_field in record.keys():
            json_clean = json_field.replace('\\n', '\n')
            for quote in quotes:
                json_clean = json_clean.replace(quote, '')

            if json_clean == template_clean:
                value = record[json_field]
                break

        if value is not None:
            mapped_record[template_col] = value
            continue

        # 5. 图片字段匹配 - 中等优先级
        if '图片' in template_col:
            for json_field in record.keys():
                if '图片' in json_field:
                    value = record[json_field]
                    break
            mapped_record[template_col] = value
            continue

        # 6. 仓库代码匹配（非FBA） - 中等优先级
        if '仓库' in template_col and '代码' in template_col:
            for json_field in record.keys():
                if '仓库' in json_field and '代码' in json_field:
                    value = record[json_field]
                    break
            mapped_record[template_col] = value
            continue

        # 7. 最后的通用模糊匹配 - 最低优先级，但对FBA字段禁用
        if 'FBA' not in template_col:
            best_match = None
            best_score = 0.8  # 要求高相似度阈值

            for json_field in record.keys():
                # 跳过FBA相关字段，避免交叉污染
                if 'FBA' in json_field:
                    continue

                # 计算相似度
                template_norm = normalize_field_name(template_col)
                json_norm = normalize_field_name(json_field)

                # 精确匹配
                if template_norm == json_norm:
                    score = len(template_norm)
                    if score > best_score and score >= 4:  # 要求至少4个字符
                        best_score = score
                        best_match = json_field
                # 长字符串包含匹配
                elif template_norm in json_norm:
                    score = len(template_norm)
                    if score > best_score and score >= 4:
                        best_score = score
                        best_match = json_field
                # 短字符串包含匹配
                elif json_norm in template_norm:
                    score = len(json_norm)
                    if score > best_score and score >= 4:
                        best_score = score
                        best_match = json_field

            if best_match:
                value = record[best_match]

        # 设置值，空值和复杂对象设为空字符串，但在设置前先处理
        if value is None or isinstance(value, (list, dict)):
            # 对于图片字段，特殊处理以确保能转换为Excel
            if '图片' in template_col and isinstance(value, list):
                value = ""  # 图片字段设为空字符串，通过单独逻辑处理
            else:
                value = ""  # 其他复杂对象也设为空字符串

        mapped_record[template_col] = value

    return mapped_record


def get_cell_size(ws, row_num: int, col_num: int) -> tuple:
    """获取单元格的实际大小（以像素为单位）"""
    try:
        # 获取列宽（以字符为单位）
        column_letter = get_column_letter(col_num)
        col_width = ws.column_dimensions[column_letter].width or 8.43  # 默认列宽

        # 获取行高（以磅为单位）
        row_height = ws.row_dimensions[row_num].height or 15  # 默认行高

        # 转换为像素
        # Excel中，1个字符宽度约等于7像素
        # 1磅约等于1.33像素
        width_px = int(col_width * 7)
        height_px = int(row_height * 1.33)

        # 减去一些边距，确保图片不会完全填满单元格
        width_px = max(width_px - 10, 20)  # 最小宽度20像素
        height_px = max(height_px - 10, 20)  # 最小高度20像素

        return width_px, height_px

    except Exception as e:
        log(f"获取单元格大小失败，使用默认值: {e}")
        return 80, 60  # 默认大小


def resize_image_to_fit_cell(img, cell_width: int, cell_height: int, padding: int = 5):
    """调整图片大小以适应单元格，保持宽高比"""
    try:
        # 获取原始图片尺寸
        original_width = img.width
        original_height = img.height

        # 计算可用空间（减去内边距）
        available_width = max(cell_width - 2 * padding, 10)
        available_height = max(cell_height - 2 * padding, 10)

        # 计算缩放比例
        width_ratio = available_width / original_width
        height_ratio = available_height / original_height

        # 选择较小的缩放比例，确保图片完全适应单元格
        scale_ratio = min(width_ratio, height_ratio, 1.0)  # 不放大图片

        # 应用缩放
        new_width = int(original_width * scale_ratio)
        new_height = int(original_height * scale_ratio)

        img.width = new_width
        img.height = new_height

    except Exception as e:
        log(f"调整图片大小失败，使用原始大小: {e}")


def process_images_in_excel(wb, ws, records: List[Dict], json_file: str) -> None:
    """处理Excel中的图片插入 - 完全参考old版本实现"""
    try:
        image_dir = resolve_image_directory(json_file)
        if not image_dir:
            return

        # 查找图片列
        image_columns = []
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value and '图片' in str(cell.value):
                image_columns.append(col_idx)

        if not image_columns:
            log("  未找到图片列")
            return

        # 插入图片
        for record_idx, record in enumerate(records):
            row_num = record_idx + 2  # 从第2行开始

            for image_col in image_columns:
                # 获取图片信息
                image_filenames = []
                if '_images' in record:
                    for img_item in record['_images']:
                        if isinstance(img_item, dict) and 'filename' in img_item:
                            image_filenames.append(img_item['filename'])

                # 为当前列插入图片（每列只插一张）
                for img_idx, img_filename in enumerate(image_filenames):
                    if img_idx >= len(image_columns):  # 防止图片数量超过图片列数
                        break

                    img_path = os.path.join(image_dir, img_filename)

                    if os.path.exists(img_path):
                        try:
                            # 加载图片
                            img = Image(img_path)

                            # 获取单元格的实际大小
                            cell_width, cell_height = get_cell_size(ws, row_num, image_col)

                            # 调整图片大小以适应单元格，保持宽高比
                            resize_image_to_fit_cell(img, cell_width, cell_height)

                            # 确定插入位置
                            col_num = image_col

                            # 获取列字母
                            if col_num <= 26:
                                col_letter = chr(64 + col_num)
                            else:
                                col_letter = chr(64 + (col_num - 1) // 26) + chr(64 + ((col_num - 1) % 26) + 1)

                            ws.add_image(img, f"{col_letter}{row_num}")
                            log(f"  插入图片: {img_filename} -> {col_letter}{row_num}")

                        except Exception as e:
                            log(f"  插入图片失败 {img_filename}: {e}")
                    else:
                        log(f"  图片文件不存在: {img_path}")

        log(f"  图片处理完成: {len(records)} 条记录")

    except Exception as e:
        log(f"图片处理失败: {e}")


def save_excel_with_template_and_images(df: pd.DataFrame, output_path: str,
                                      records: List[Dict], json_file: str,
                                      template_path: str = None) -> None:
    """使用模板格式保存Excel文件并插入图片 - 完全参考old版本"""
    try:
        if template_path is None:
            template_path = DEFAULT_TEMPLATE

        # 确保输出目录存在
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # 如果目标文件已存在，先删除
        if os.path.exists(output_path):
            try:
                os.remove(output_path)
            except PermissionError:
                log(f"文件被占用，无法删除: {output_path}")
                return

        # 复制模板文件 - 这是保持格式的关键
        shutil.copy2(template_path, output_path)

        # 使用openpyxl加载
        wb = load_workbook(output_path)
        ws = wb.active

        # 设置行高以容纳图片
        ws.row_dimensions[1].height = 40  # 标题行

        # 从第二行开始写入数据
        start_row = 2

        # 写入数据行
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

            # 设置适当的行高
            ws.row_dimensions[r_idx].height = 80

        # 插入图片
        process_images_in_excel(wb, ws, records, json_file)

        # 保持模板原有格式，但确保数据行使用标准文本格式
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            for cell in row:
                if cell.value is not None and row_idx > 1:  # 只修改数据行，保持标题行格式
                    # 获取原始对齐方式
                    original_horizontal = cell.alignment.horizontal if cell.alignment else 'center'

                    # 设置标准字体格式（黑色、无下划线）
                    cell.font = Font(
                        name='Calibri',
                        size=11,
                        bold=False,  # 数据行不使用粗体
                        italic=False,
                        color='000000',  # 黑色
                        underline='none'  # 无下划线
                    )

                    # 设置对齐方式
                    cell.alignment = Alignment(
                        horizontal=original_horizontal,
                        vertical='top',
                        wrap_text=True
                    )

        # 保存文件
        wb.save(output_path)
        log(f"使用模板格式保存成功，图片已插入: {os.path.basename(output_path)}")

    except Exception as e:
        log(f"保存Excel失败: {e}")


def json_to_excel(input_path: str, output_path: str, template_path: str = None, process_images: bool = True) -> bool:
    """JSON转Excel主函数"""
    try:
        # 设置默认模板
        if template_path is None:
            template_path = DEFAULT_TEMPLATE

        log(f"开始转换: {os.path.basename(input_path)}")

        # 加载JSON数据
        json_data = load_json_data(input_path)
        records = json_data.get('data', [])

        if not records:
            log("没有数据记录")
            return False

        # 获取模板列
        template_columns = find_template_columns(template_path)
        if not template_columns:
            log("模板列获取失败")
            return False

        # 映射数据到模板
        mapped_data = []
        for record in records:
            mapped_record = map_fields_to_template(record, template_columns)
            mapped_data.append(mapped_record)

        # 再次清理所有复杂数据类型，确保能转换为Excel
        # 同时统一数值格式处理，避免不同文件格式差异
        for record in mapped_data:
            for key, value in record.items():
                if isinstance(value, (list, dict)):
                    record[key] = ""
                elif value is None:
                    record[key] = ""
                elif isinstance(value, str):
                    # 尝试转换字符串为数值，统一格式
                    try:
                        # 检查是否为数字字符串
                        if value.replace('.', '', 1).replace('-', '', 1).isdigit():
                            # 如果有小数点，转换为float；否则转换为int
                            if '.' in value:
                                record[key] = float(value)
                            else:
                                record[key] = int(value)
                        else:
                            # 保持字符串不变
                            pass
                    except ValueError:
                        # 转换失败，保持原字符串
                        pass

        # 创建DataFrame
        df = pd.DataFrame(mapped_data)

        # 确保输出目录存在
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # 保存Excel
        if process_images:
            # 使用模板并插入图片
            save_excel_with_template_and_images(df, output_path, records, input_path, template_path)
        else:
            # 直接保存DataFrame
            df.to_excel(output_path, index=False)
            log(f"保存完成: {os.path.basename(output_path)}")

        return True

    except Exception as e:
        log(f"转换失败: {e}")
        return False


def batch_convert_json(input_dir: str, output_dir: str, template_path: str = None, process_images: bool = True) -> dict:
    """批量转换JSON文件 - 修复文件扫描逻辑"""
    results = {'success': [], 'failed': []}

    try:
        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)

        log(f"批量转换开始: {input_dir}")

        # 只扫描实际存在的子目录
        input_path = Path(input_dir)
        if not input_path.exists():
            log(f"输入目录不存在: {input_dir}")
            return results

        # 获取子目录
        subdirs = [d for d in input_path.iterdir() if d.is_dir()]

        if not subdirs:
            log("未找到子目录")
            return results

        # 转换每个子目录中的JSON文件
        for subdir in subdirs:
            # 查找子目录中的JSON文件
            json_files = list(subdir.glob("*.json"))

            if not json_files:
                continue

            # 转换每个JSON文件（通常只有一个主文件）
            for json_file in json_files:
                try:
                    # 生成输出文件名 - 使用目录名
                    company_name = subdir.name
                    output_file = os.path.join(output_dir, f"{company_name}.xlsx")

                    success = json_to_excel(str(json_file), output_file, template_path, process_images)

                    if success:
                        results['success'].append(str(json_file))
                        log(f"  成功: {subdir.name}/{json_file.name}")
                    else:
                        results['failed'].append(str(json_file))
                        log(f"  失败: {subdir.name}/{json_file.name}")

                except Exception as e:
                    results['failed'].append(str(json_file))
                    log(f"  失败: {subdir.name}/{json_file.name} - {e}")

        log(f"批量转换完成: 成功 {len(results['success'])}, 失败 {len(results['failed'])}")
        return results

    except Exception as e:
        log(f"批量转换失败: {e}")
        return results


def main():
    """命令行入口 - 与其他模块保持一致的参数结构"""
    parser = argparse.ArgumentParser(description="JSON转Excel转换器")
    parser.add_argument("--input", "-i", default="data/transformed", help="输入目录或文件")
    parser.add_argument("--output", "-o", default="data/output", help="输出目录")
    parser.add_argument("--file", "-f", help="单个文件路径")
    parser.add_argument("--template", "-t", default=DEFAULT_TEMPLATE, help="模板文件")
    parser.add_argument("--no-images", action="store_true", help="跳过图片处理")
    parser.add_argument("--overwrite", action="store_true", help="覆盖已存在文件")

    args = parser.parse_args()

    if args.file:
        # 单文件转换
        success = json_to_excel(args.file, args.output, args.template, not args.no_images)
        if success:
            print(f"转换成功: {args.file} -> {args.output}")
        else:
            print(f"转换失败: {args.file}")
            sys.exit(1)
    else:
        # 批量转换
        results = batch_convert_json(args.input, args.output, args.template, not args.no_images)
        if results['success']:
            print(f"批量转换完成，成功: {len(results['success'])}, 失败: {len(results['failed'])}")
        else:
            print("未找到可转换的文件")
            sys.exit(1)


if __name__ == "__main__":
    main()